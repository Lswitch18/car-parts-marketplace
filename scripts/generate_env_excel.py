#!/usr/bin/env python3
"""
Digital AIGarage (DAIG) - Gerador de Planilha de Configurações (.env)
Gera uma planilha Excel estilizada e profissional documentando todas as variáveis de ambiente.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ENV_PATH = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".env"))
OUTPUT_EXCEL = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "docs", "diag", "env_configuracoes.xlsx"))
os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)

def parse_env_file(filepath):
    variables = []
    if not os.path.exists(filepath):
        print(f"❌ Arquivo .env não encontrado em {filepath}")
        return variables

    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()

    current_category = "Geral"
    for line in lines:
        line_str = line.strip()
        if not line_str:
            continue
        if line_str.startswith("#"):
            # Atualiza a categoria baseada no comentário
            comment = line_str.lstrip("#").strip()
            if comment and not comment.startswith("="):
                current_category = comment
            continue

        if "=" in line_str:
            key, val = line_str.split("=", 1)
            key = key.strip()
            val = val.strip().strip('"').strip("'")
            
            # Definir escopo (Frontend vs Backend)
            scope = "Frontend (Público)" if key.startswith("VITE_") else "Backend (Servidor/Secret)"

            # Mascarar valores sensíveis
            if any(s in key.upper() for s in ["KEY", "SECRET", "TOKEN", "PASSWORD", "AUTH"]):
                if len(val) > 8:
                    masked_val = val[:4] + "••••••••" + val[-4:]
                else:
                    masked_val = "••••••••"
                is_secret = "Sim (Protegido)"
            else:
                masked_val = val
                is_secret = "Não"

            variables.append({
                "categoria": current_category,
                "chave": key,
                "valor_mascarado": masked_val,
                "escopo": scope,
                "segredo": is_secret,
                "status": "Configurado" if val else "Pendente"
            })

    return variables

def build_excel_report(variables, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Configurações DAIG"
    ws.views.sheetView[0].showGridLines = True

    # Estilos
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    fill_title = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")

    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="0EA5E9", end_color="0EA5E9", fill_type="solid")

    font_data = Font(name="Calibri", size=10, color="1F2937")
    font_bold = Font(name="Calibri", size=10, bold=True, color="1F2937")

    border_thin = Side(border_style="thin", color="E5E7EB")
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    # Título
    ws.merge_cells("A1:F1")
    c_title = ws["A1"]
    c_title.value = "DIGITAL AIGARAGE (DAIG) - MAPA DE VARIÁVEIS DE AMBIENTE (.ENV)"
    c_title.font = font_title
    c_title.fill = fill_title
    c_title.alignment = align_center
    ws.row_dimensions[1].height = 40

    # Subtítulo
    ws.merge_cells("A2:F2")
    ws["A2"].value = "Documento auditado da infraestrutura de APIs, Chaves Stripe (Japão), Supabase, Redis e LLMs."
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="6B7280")
    ws["A2"].alignment = align_left
    ws.row_dimensions[2].height = 20

    # Cabeçalhos
    headers = ["Categoria / Serviço", "Nome da Variável", "Valor Mascarado / Referência", "Escopo de Execução", "Dado Sensível?", "Status"]
    ws.append([]) # linha 3 vazia

    ws.row_dimensions[4].height = 28
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = box_border

    start_row = 5
    for idx, var in enumerate(variables, start=start_row):
        ws.row_dimensions[idx].height = 22
        
        ws.cell(row=idx, column=1, value=var["categoria"]).alignment = align_left
        ws.cell(row=idx, column=2, value=var["chave"]).font = font_bold
        ws.cell(row=idx, column=3, value=var["valor_mascarado"])
        ws.cell(row=idx, column=4, value=var["escopo"]).alignment = align_center
        ws.cell(row=idx, column=5, value=var["segredo"]).alignment = align_center
        
        c_status = ws.cell(row=idx, column=6, value=var["status"])
        c_status.alignment = align_center
        if var["status"] == "Configurado":
            c_status.fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
            c_status.font = Font(name="Calibri", size=10, bold=True, color="166534")

        for c in range(1, 7):
            ws.cell(row=idx, column=c).border = box_border
            if c != 2 and c != 6:
                ws.cell(row=idx, column=c).font = font_data

    # Ajuste de largura das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save(output_path)
    print(f"✅ Planilha .env gerada com sucesso em: {output_path}")

if __name__ == "__main__":
    vars_list = parse_env_file(ENV_PATH)
    build_excel_report(vars_list, OUTPUT_EXCEL)
